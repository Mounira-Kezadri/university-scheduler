"""
================================================================================
نظام تصميم وإدارة الجداول الدراسية الجامعية
University Academic Schedule Design & Management System
--------------------------------------------------------------------------------
Developer : Dr. Mounira Kezadri
Institution: Applied College — Taibah University
Unit       : Computer Science & Information Programs
Version    : 1.0
================================================================================
"""

import streamlit as st
import pandas as pd
import io
import re
import json
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
 
# ── Logging setup ─────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)
 
# ══════════════════════════════════════════════════════════════════════════════
# 1.  CONFIGURATION  (centralised — easy to extend for other institutions)
# ══════════════════════════════════════════════════════════════════════════════
 
# Default time slots — can be overridden via config.json in the same directory
DEFAULT_SLOTS = [
    {"label": "8:30-9:45",   "start_min": 8*60+30, "end_min": 9*60+45},
    {"label": "10:00-11:15", "start_min": 10*60,   "end_min": 11*60+15},
    {"label": "11:30-12:45", "start_min": 11*60+30, "end_min": 12*60+45},
    {"label": "13:00-14:15", "start_min": 13*60,   "end_min": 14*60+15},
]
 
def load_config() -> list:
    """Load time slots from config.json if present, otherwise use defaults."""
    try:
        with open("config.json", "r", encoding="utf-8") as fh:
            cfg = json.load(fh)
            slots = cfg.get("time_slots", DEFAULT_SLOTS)
            logger.info("Loaded %d slots from config.json", len(slots))
            return slots
    except FileNotFoundError:
        return DEFAULT_SLOTS
    except (json.JSONDecodeError, KeyError) as exc:
        logger.warning("config.json error (%s) — using defaults.", exc)
        return DEFAULT_SLOTS
 
STANDARD_SLOTS: list = load_config()
 
DAYS_EN   = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
DAYS_MAP  = {
    'Sunday': 'الاحد', 'Monday': 'الاثنين', 'Tuesday': 'الثلاثاء',
    'Wednesday': 'الأربعاء', 'Thursday': 'الخميس'
}
INVALID_ROOMS = {"?", "nan", "", "N/A"}
 
# ── Excel style constants ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
EMPTY_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
 
# ══════════════════════════════════════════════════════════════════════════════
# 2.  PARSING UTILITIES
# ══════════════════════════════════════════════════════════════════════════════
 
def split_course_code(code: str) -> tuple[str, str]:
    """Split 'CS101' → ('CS', '101').  Returns (code, '') if no numeric suffix."""
    match = re.match(r"([a-zA-Z]+)([0-9]+)", str(code))
    if match:
        return match.group(1), match.group(2)
    return str(code), ""
 
 
def parse_time_range(t_str: str) -> tuple[int, int]:
    """
    Convert 'HH:MM-HH:MM' or 'HH:MM–HH:MM' to (start_minutes, end_minutes).
    Returns (0, 0) on failure and logs a warning.
    """
    cleaned = str(t_str).replace('\u2013', '-').replace('\u2014', '-').strip()
    parts = cleaned.split('-')
    if len(parts) != 2:
        logger.warning("Unrecognised time format: '%s'", t_str)
        return 0, 0
    try:
        s = datetime.strptime(parts[0].strip(), "%H:%M")
        e = datetime.strptime(parts[1].strip(), "%H:%M")
        return s.hour * 60 + s.minute, e.hour * 60 + e.minute
    except ValueError:
        logger.warning("Cannot parse time value: '%s'", t_str)
        return 0, 0
 
 
def _detect_row_pattern(df: pd.DataFrame, r: int, c: int) -> int:
    """
    Detect whether data uses 3-row or 4-row layout for a given cell.
 
    4-row layout (row offset):  0=day, 1=course_name, 2=course_code, 3=time, 4=room
    3-row layout (row offset):  0=day, 1=course_code, 2=time, 3=room
 
    Returns the name_offset (1 for 4-row, 0 for 3-row — 0 means no name row).
    """
    try:
        candidate = str(df.iloc[r + 1, c]).strip()
        # If the cell after the day row looks like a course code (letters+digits)
        # then we are in 3-row mode (no Arabic name row).
        if re.match(r"^[a-zA-Z]+\d+$", candidate):
            return 0   # 3-row: code is at r+1
        return 1       # 4-row: name is at r+1, code is at r+2 — but code was already at r
    except IndexError:
        return 1
 
 
def parse_excel_files(files) -> pd.DataFrame:
    """
    Read one or more uploaded Excel files and return a unified DataFrame.
 
    Supports both 3-row and 4-row per-course layouts, detected automatically
    per cell.  Handles both regular hyphen and em-dash time separators.
    """
    all_data: list[dict] = []
 
    for f in files:
        # Read file bytes once — avoid exhausting the stream on re-reads
        raw = io.BytesIO(f.read())
        try:
            xl = pd.ExcelFile(raw)
        except Exception as exc:
            st.warning(f"⚠️  Could not open '{f.name}': {exc}")
            continue
 
        for sheet in xl.sheet_names:
            try:
                df = xl.parse(sheet, header=None).reset_index(drop=True)
            except Exception as exc:
                st.warning(f"⚠️  Could not parse sheet '{sheet}' in '{f.name}': {exc}")
                continue
 
            section = "Unknown"
 
            for r in range(len(df) - 1):          # -1 guards against out-of-bounds
                val = str(df.iloc[r, 0]).strip()
 
                # ── Section header (e.g. F1, F2, …) ──────────────────────────
                if val and val != "nan" and not val.isdigit() and val not in DAYS_MAP:
                    section = val
                    continue
 
                if val not in DAYS_MAP:
                    continue
 
                day_en = val
 
                for c in range(1, df.shape[1]):
                    code_full = str(df.iloc[r, c]).strip()
 
                    # Skip empty / numeric-only cells
                    if not code_full or code_full.lower() in {"nan", "1", "2", "3", "4", "5"}:
                        continue
 
                    # ── Detect 3-row vs 4-row layout per cell ─────────────────
                    name_offset = _detect_row_pattern(df, r, c)
 
                    try:
                        if name_offset == 1:
                            # 4-row: day | name | code | time | room
                            course_name = str(df.iloc[r + 1, c]).strip()
                            time_str    = str(df.iloc[r + 2, c]).strip()
                            room        = str(df.iloc[r + 3, c]).strip()
                        else:
                            # 3-row: day | code | time | room
                            course_name = ""
                            time_str    = str(df.iloc[r + 1, c]).strip()
                            room        = str(df.iloc[r + 2, c]).strip()
                    except IndexError:
                        logger.warning("Row index out of range at sheet=%s r=%d c=%d", sheet, r, c)
                        continue
 
                    prefix, num = split_course_code(code_full)
                    start_m, end_m = parse_time_range(time_str)
 
                    # Skip rows where time could not be parsed (both values stay 0)
                    if start_m == 0 and end_m == 0:
                        continue
 
                    all_data.append({
                        "التخصص":  sheet,
                        "المقرر":  course_name,
                        "الرمز":   prefix,
                        "الرقم":   num,
                        "الشعبة":  section,
                        "Day":     day_en,
                        "Time":    time_str,
                        "القاعة":  room,
                        "StartMin": start_m,
                        "EndMin":   end_m,
                    })
 
    if not all_data:
        return pd.DataFrame()
 
    return pd.DataFrame(all_data)
 
 
# ══════════════════════════════════════════════════════════════════════════════
# 3.  CONFLICT DETECTION
# ══════════════════════════════════════════════════════════════════════════════
 
def detect_conflicts(df: pd.DataFrame) -> list[dict]:
    """
    Return a list of room-time conflicts (two different courses/sections
    assigned to the same room at overlapping times on the same day).
    """
    if df.empty:
        return []
 
    conflicts: list[dict] = []
    valid_df = df[~df['القاعة'].isin(INVALID_ROOMS)].copy()
    valid_df = valid_df.sort_values(['القاعة', 'Day', 'StartMin'])
 
    for (room, day), group in valid_df.groupby(['القاعة', 'Day']):
        rows = group.to_dict('records')
        for i in range(len(rows) - 1):
            curr, nxt = rows[i], rows[i + 1]
            overlap = nxt['StartMin'] < curr['EndMin']
            same_entry = (curr['الرمز'] == nxt['الرمز'] and curr['الشعبة'] == nxt['الشعبة'])
            if overlap and not same_entry:
                conflicts.append({
                    "القاعة":    room,
                    "اليوم":     DAYS_MAP.get(day, day),
                    "التفاصيل":  f"تداخل بين [{curr['الرمز']}{curr['الرقم']}] و [{nxt['الرمز']}{nxt['الرقم']}]",
                    "الشعبتان":  f"{curr['الشعبة']} / {nxt['الشعبة']}",
                    "الوقت":     f"{curr['Time']} ↔ {nxt['Time']}",
                })
    return conflicts
 
 
# ══════════════════════════════════════════════════════════════════════════════
# 4.  EXPORT FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
 
def _header_row(ws, values: list, col_widths: dict | None = None):
    """Write a styled header row to a worksheet."""
    ws.append(values)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = BORDER
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
 
 
def export_linear_format(df: pd.DataFrame) -> bytes:
    """
    Export a flat linear schedule sorted by specialization → course → section.
    Each row = one section; day columns contain the time slot for that day.
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Linear Schedule"
    ws.sheet_view.rightToLeft = True
 
    headers = ["التخصص", "المقرر", "الرمز", "الرقم", "الشعبة",
               "الاحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "القاعة"]
    col_widths = {'A': 15, 'B': 28, 'C': 10, 'D': 8, 'E': 10,
                  'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 16, 'K': 15}
    _header_row(ws, headers, col_widths)
 
    days_cols = {'Sunday': 6, 'Monday': 7, 'Tuesday': 8, 'Wednesday': 9, 'Thursday': 10}
    df_sorted = df.sort_values(['التخصص', 'المقرر', 'الشعبة'])
    grouped   = df_sorted.groupby(
        ['التخصص', 'المقرر', 'الرمز', 'الرقم', 'الشعبة', 'القاعة'], sort=False
    )
 
    last_item_id = None
    use_gray     = False
 
    for (spec, name, pref, num, sec, room), group in grouped:
        current_id = (spec, name, sec)
        if current_id != last_item_id:
            use_gray     = not use_gray
            last_item_id = current_id
 
        row_data = [spec, name, pref, num, sec, "", "", "", "", "", room]
        for _, item in group.iterrows():
            col_idx = days_cols.get(item['Day'])
            if col_idx:
                row_data[col_idx - 1] = item['Time']
 
        ws.append(row_data)
        fill = GRAY_FILL if use_gray else WHITE_FILL
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
 
    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()
 
 
def export_rooms_styled(df: pd.DataFrame, mode: str = "occupied") -> bytes:
    """
    Export room schedules.
    mode='occupied' → one sheet per room, weekly grid with color-coded empty slots.
    mode='empty'    → single sheet listing free rooms per day/slot.
    """
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    slots = [s["label"] for s in STANDARD_SLOTS]
 
    if mode == "occupied":
        valid_rooms = sorted(
            r for r in df['القاعة'].unique() if r not in INVALID_ROOMS
        )
        for room in valid_rooms:
            ws = wb.create_sheet(title=str(room)[:31])
            ws.sheet_view.rightToLeft = True
            col_widths = {'A': 15}
            col_widths.update({get_column_letter(i): 25 for i in range(2, len(slots) + 2)})
            _header_row(ws, ["اليوم / الحصة"] + slots, col_widths)
 
            room_df = df[df['القاعة'] == room]
 
            for d_en in DAYS_EN:
                row_vals = [DAYS_MAP.get(d_en, d_en)]
                day_df   = room_df[room_df['Day'] == d_en]
 
                for slot_cfg in STANDARD_SLOTS:
                    cell_lines: list[str] = []
                    for _, row in day_df.iterrows():
                        if abs(row['StartMin'] - slot_cfg["start_min"]) <= 30:
                            cell_lines.append(
                                f"{row['الرمز']}{row['الرقم']}\n{row['المقرر']}\n({row['الشعبة']})"
                            )
                    row_vals.append("\n\n".join(cell_lines))
 
                ws.append(row_vals)
 
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = 70
                for cell in row:
                    cell.border    = BORDER
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    if cell.column > 1 and not cell.value:
                        cell.fill = EMPTY_FILL
 
            ws.freeze_panes = "B2"
 
    else:
        # ── Available rooms sheet ──────────────────────────────────────────────
        ws = wb.create_sheet(title="Available Rooms")
        ws.sheet_view.rightToLeft = True
        col_widths = {'A': 16}
        col_widths.update({get_column_letter(i): 35 for i in range(2, len(slots) + 2)})
        _header_row(ws, ["Day / Slot"] + slots, col_widths)
 
        all_rooms = set(df['القاعة'].unique()) - INVALID_ROOMS
 
        for d_en in DAYS_EN:
            row_data = [DAYS_MAP.get(d_en, d_en)]
            day_df   = df[df['Day'] == d_en]
 
            for slot_cfg in STANDARD_SLOTS:
                s_min = slot_cfg["start_min"]   # capture value, not reference
                mask  = day_df['StartMin'].apply(lambda x, s=s_min: abs(x - s) <= 30)
                occupied = set(day_df.loc[mask, 'القاعة'].unique()) - INVALID_ROOMS
                free = sorted(all_rooms - occupied)
                row_data.append(", ".join(free))
 
            ws.append(row_data)
 
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 55
            for cell in row:
                cell.border    = BORDER
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
 
        ws.freeze_panes = "B2"
 
    wb.save(output)
    return output.getvalue()
 
 
# ══════════════════════════════════════════════════════════════════════════════
# 5.  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════
 
def render_metrics(data: pd.DataFrame, conflicts: list):
    """Show a quick-stats bar at the top of the page."""
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📚 المقررات",  (data['الرمز'] + data['الرقم']).nunique() if not data.empty else 0)
    c2.metric("🏢 القاعات",   data['القاعة'].nunique()   if not data.empty else 0)
    c3.metric("👥 الشعب",     data['الشعبة'].nunique()   if not data.empty else 0)
    c4.metric(
        "⚠️ التعارضات", len(conflicts),
        delta="تعارض" if conflicts else None,
        delta_color="inverse"
    )
 
 
def main():
    st.set_page_config(
        page_title="نظام تصميم وإدارة الجداول",
        page_icon="📅",
        layout="wide",
    )
 
    # ── Page header ────────────────────────────────────────────────────────────
    st.title("📅 نظام تصميم وإدارة الجداول الدراسية")
    st.caption("المطور: د. منيرة قزادري - الكلية التطبيقية - جامعة طيبة")
    st.divider()
 
    # ── File upload ────────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "ارفع ملفات Excel (ملف لكل تخصص)",
        type="xlsx",
        accept_multiple_files=True,
        help="يجب أن تكون الملفات بصيغة .xlsx وتتبع الهيكل الموحَّد الموضَّح في README",
    )
 
    if not uploaded:
        st.info("⬆️  الرجاء رفع ملف اكسل أو أكثر للبدء ")
        return
 
    # ── Parse & cache in session_state (avoids re-parsing on every widget click)
    file_key = tuple(f.name + str(f.size) for f in uploaded)
 
    if st.session_state.get("file_key") != file_key:
        with st.spinner("⏳ جارٍ تحليل الملفات…"):
            data      = parse_excel_files(uploaded)
            conflicts = detect_conflicts(data)
        st.session_state["file_key"]  = file_key
        st.session_state["data"]      = data
        st.session_state["conflicts"] = conflicts
 
    data      = st.session_state["data"]
    conflicts = st.session_state["conflicts"]
 
    if data.empty:
        st.error("❌ لم يُعثَر على بيانات صالحة. تحققي من هيكل الملفات.")
        return
 
    # ── Quick-stats bar ────────────────────────────────────────────────────────
    render_metrics(data, conflicts)
    st.divider()
 
    # ══════════════════════════════════════════════════════════════════════════
    # TABS
    # ══════════════════════════════════════════════════════════════════════════
    tab1, tab2, tab3 = st.tabs([
        "📋 التنسيق الخطي",
        "🏢 القاعات والشواغر",
        "⚠️ التعارضات",
    ])
 
    # ── Tab 1 : Linear schedule ────────────────────────────────────────────────
    with tab1:
        st.subheader("المعاينة الخطية — مرتبة حسب التخصص")
 
        col_filter, col_dl = st.columns([3, 1])
        with col_filter:
            specs = ["الكل"] + sorted(data['التخصص'].unique().tolist())
            chosen = st.selectbox("🔎 تصفية حسب التخصص", specs)
        with col_dl:
            st.write("")          # vertical alignment spacer
            st.download_button(
                "📥 تحميل Excel",
                export_linear_format(data),
                "Linear_Clean_Schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
 
        display_df = data if chosen == "الكل" else data[data['التخصص'] == chosen]
        st.dataframe(
            display_df.sort_values(["التخصص", "المقرر"])
                      [["التخصص", "المقرر", "الرمز", "الرقم", "الشعبة", "Day", "Time", "القاعة"]],
            use_container_width=True,
            height=500,
        )
 
    # ── Tab 2 : Rooms ──────────────────────────────────────────────────────────
    with tab2:
        st.subheader("جداول القاعات والشواغر")
 
        # Room preview inside the app
        valid_rooms = sorted(r for r in data['القاعة'].unique() if r not in INVALID_ROOMS)
        selected_room = st.selectbox("👁️ معاينة جدول قاعة", valid_rooms)
 
        if selected_room:
            room_df = data[data['القاعة'] == selected_room]
            slots   = [s["label"] for s in STANDARD_SLOTS]
            preview = []
            for d_en in DAYS_EN:
                row_dict = {"اليوم": DAYS_MAP.get(d_en, d_en)}
                day_df   = room_df[room_df['Day'] == d_en]
                for slot_cfg in STANDARD_SLOTS:
                    s_min   = slot_cfg["start_min"]   # capture value, not reference
                    matches = day_df[day_df['StartMin'].apply(lambda x, s=s_min: abs(x - s) <= 30)]
                    if matches.empty:
                        row_dict[slot_cfg["label"]] = "🟢 فارغة"
                    else:
                        row_dict[slot_cfg["label"]] = " | ".join(
                            f"{r['الرمز']}{r['الرقم']} ({r['الشعبة']})"
                            for _, r in matches.iterrows()
                        )
                preview.append(row_dict)
 
            st.dataframe(pd.DataFrame(preview).set_index("اليوم"),
                         use_container_width=True)
 
        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "🏢 تحميل جدول القاعات (ملون)",
                export_rooms_styled(data, "occupied"),
                "Rooms_Schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col2:
            st.download_button(
                "✅ تحميل القاعات الفارغة",
                export_rooms_styled(data, "empty"),
                "Available_Rooms.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
 
    # ── Tab 3 : Conflicts ──────────────────────────────────────────────────────
    with tab3:
        st.subheader("تقرير التعارضات الزمنية")
        if conflicts:
            st.error(f"⚠️  يوجد **{len(conflicts)}** تداخل في القاعات — يرجى المراجعة قبل الاعتماد.")
            conflict_df = pd.DataFrame(conflicts)
            # Allow filtering by room
            rooms_in_conflict = ["الكل"] + sorted(conflict_df['القاعة'].unique().tolist())
            chosen_room = st.selectbox("🔎 تصفية حسب القاعة", rooms_in_conflict)
            if chosen_room != "الكل":
                conflict_df = conflict_df[conflict_df['القاعة'] == chosen_room]
            st.dataframe(conflict_df, use_container_width=True)
        else:
            st.success("✅ لا توجد تداخلات زمنية — الجداول سليمة.")
 
 
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    main()
